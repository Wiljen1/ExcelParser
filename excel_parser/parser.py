import json
from openpyxl import load_workbook
from .logic import CONFIG


def build_merged_lookup(ws):
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(r, c)] = value
    return lookup


def get_value(ws, lookup, r, c):
    val = ws.cell(row=r, column=c).value
    if val is not None:
        return val
    return lookup.get((r, c))


def normalize(val):
    if val is None:
        return "NO"
    if str(val).strip() == "X":
        return "YES"
    return val


def parse_file(path, sheet_name, output_path):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    lookup = build_merged_lookup(ws)

    results = []

    for col in range(CONFIG["start_col"], CONFIG["end_col"] + 1):
        suite = get_value(ws, lookup, 6, col)
        version = get_value(ws, lookup, 5, col)

        for row in range(CONFIG["data_start_row"], CONFIG["data_end_row"] + 1):
            field = ws.cell(row=row, column=2).value
            if not field:
                continue

            value = normalize(get_value(ws, lookup, row, col))

            results.append({
                "suite": suite,
                "version": version,
                "field": field,
                "value": value,
                "row": row,
                "col": col
            })

    with open(output_path, "w") as f:
        json.dump(results, f, indent=2)

    return results
