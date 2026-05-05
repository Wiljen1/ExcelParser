import json
from openpyxl import load_workbook
from .logic_v2 import CONFIG


def build_merged_lookup(ws):
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(r, c)] = value
    return lookup


def get_value(ws, lookup, r, c):
    val = ws.cell(row=r, column=c).value
    if val is not None:
        return val
    return lookup.get((r, c))


def get_hyperlink(ws, r, c):
    cell = ws.cell(row=r, column=c)
    if cell.hyperlink:
        return {"title": cell.value, "url": cell.hyperlink.target}
    return None


def normalize(val):
    if val is None:
        return "NO"
    if str(val).strip().upper() == "X":
        return "YES"
    return val


def get_comment(value):
    if value is None:
        return None
    text = str(value)
    for marker, comment in CONFIG.get("footnote_comments", {}).items():
        if marker in text:
            return comment
    return None


def parse_file(path, sheet_name, output_path):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    lookup = build_merged_lookup(ws)
    records = []

    for col in range(CONFIG["start_col"], CONFIG["end_col"] + 1):
        suite = get_value(ws, lookup, 6, col)
        version = get_value(ws, lookup, 5, col)

        for group in CONFIG["groups"]:
            title = ws.cell(row=group["title_row"], column=2).value

            for row in range(group["child_start"], group["child_end"] + 1):
                field = ws.cell(row=row, column=2).value
                if not field:
                    continue

                raw_value = get_value(ws, lookup, row, col)
                value = normalize(raw_value)
                link = get_hyperlink(ws, row, col)
                comment = get_comment(raw_value)

                record = {
                    "suite": suite,
                    "version": version,
                    "title": title,
                    "field": field,
                    "value": value,
                    "row": row,
                    "col": col
                }

                if link:
                    record["link"] = link
                if comment:
                    record["comment"] = comment

                records.append(record)

    output = {
        "sheet": sheet_name,
        "records": records,
        "additional_information": CONFIG.get("additional_information", [])
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    return output
