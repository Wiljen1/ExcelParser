import json
import re
from openpyxl import load_workbook
from .logic_v2 import CONFIG

URL_RE = re.compile(r"https?://[^\s\)\]\}\>\"']+")
HYPERLINK_FORMULA_RE = re.compile(r'=HYPERLINK\("([^"]+)"\s*,\s*"?([^"]*)"?\)', re.IGNORECASE)


def build_merged_lookup(ws):
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(r, c)] = value
    return lookup


def build_merged_coordinate_lookup(ws):
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = (merged_range.min_row, merged_range.min_col)
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(r, c)] = top_left
    return lookup


def get_value(ws, value_lookup, r, c):
    val = ws.cell(row=r, column=c).value
    if val is not None:
        return val
    return value_lookup.get((r, c))


def get_display_title(value):
    if value is None:
        return None
    return str(value)


def extract_links_from_text(value):
    if value is None:
        return []
    text = str(value)
    return [{"title": text, "url": url} for url in URL_RE.findall(text)]


def extract_hyperlink_formula(value):
    if value is None:
        return None
    text = str(value).strip()
    match = HYPERLINK_FORMULA_RE.match(text)
    if not match:
        return None
    url = match.group(1)
    title = match.group(2) or url
    return {"title": title, "url": url}


def get_hyperlinks(ws_values, ws_formulas, merged_coord_lookup, r, c, display_value):
    links = []

    # 1. Check the exact cell and the top-left cell of a merged range.
    coordinates_to_check = [(r, c)]
    merged_top_left = merged_coord_lookup.get((r, c))
    if merged_top_left and merged_top_left not in coordinates_to_check:
        coordinates_to_check.append(merged_top_left)

    for check_r, check_c in coordinates_to_check:
        for ws in (ws_values, ws_formulas):
            cell = ws.cell(row=check_r, column=check_c)
            if cell.hyperlink and cell.hyperlink.target:
                link = {
                    "title": get_display_title(display_value) or get_display_title(cell.value) or cell.hyperlink.target,
                    "url": cell.hyperlink.target
                }
                if link not in links:
                    links.append(link)

    # 2. Check if the cell is an Excel HYPERLINK formula.
    formula_value = get_value(ws_formulas, {}, r, c)
    formula_link = extract_hyperlink_formula(formula_value)
    if formula_link and formula_link not in links:
        links.append(formula_link)

    if merged_top_left:
        formula_top_left = get_value(ws_formulas, {}, merged_top_left[0], merged_top_left[1])
        formula_link = extract_hyperlink_formula(formula_top_left)
        if formula_link and formula_link not in links:
            links.append(formula_link)

    # 3. Check if the visible value itself contains a plain URL.
    for text_link in extract_links_from_text(display_value):
        if text_link not in links:
            links.append(text_link)

    return links


def normalize(val):
    if val is None:
        return "NO"
    if str(val).strip().upper() == "X":
        return "YES"
    return val


def get_comment(value):
    if value is None:
        return None
    text = str(value)
    for marker, comment in CONFIG.get("footnote_comments", {}).items():
        if marker in text:
            return comment
    return None


def parse_file(path, sheet_name, output_path):
    # values workbook gives calculated/displayed values.
    wb_values = load_workbook(path, data_only=True)
    ws_values = wb_values[sheet_name]

    # formulas workbook preserves =HYPERLINK(...) formulas.
    wb_formulas = load_workbook(path, data_only=False)
    ws_formulas = wb_formulas[sheet_name]

    value_lookup = build_merged_lookup(ws_values)
    merged_coord_lookup = build_merged_coordinate_lookup(ws_values)
    records = []

    for col in range(CONFIG["start_col"], CONFIG["end_col"] + 1):
        suite = get_value(ws_values, value_lookup, 6, col)
        version = get_value(ws_values, value_lookup, 5, col)

        for group in CONFIG["groups"]:
            title = ws_values.cell(row=group["title_row"], column=2).value

            for row in range(group["child_start"], group["child_end"] + 1):
                field = ws_values.cell(row=row, column=2).value
                if not field:
                    continue

                raw_value = get_value(ws_values, value_lookup, row, col)
                value = normalize(raw_value)
                links = get_hyperlinks(ws_values, ws_formulas, merged_coord_lookup, row, col, raw_value)
                comment = get_comment(raw_value)

                record = {
                    "suite": suite,
                    "version": version,
                    "title": title,
                    "field": field,
                    "value": value,
                    "row": row,
                    "col": col,
                    "source_cell": ws_values.cell(row=row, column=col).coordinate
                }

                if links:
                    record["links"] = links
                    # Keep a single legacy field too, for easy consumption.
                    record["link"] = links[0]

                if comment:
                    record["comment"] = comment

                records.append(record)

    output = {
        "sheet": sheet_name,
        "records": records,
        "additional_information": CONFIG.get("additional_information", [])
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    return output
